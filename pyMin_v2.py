'''
Created on Mar 28, 2019
first stable version April 17 2019

@author: miki
'''

##SYSTEM
import os
import sys

from tkinter import *
from tkinter import filedialog
from tkinter import scrolledtext
#import xlsxwriter
import openpyxl
import pandas as pd
import PySimpleGUI as sg
import seaborn as sns
from pandasgui import show
#import matplotlib
#matplotlib.use("TkAgg")
# from matplotlib import interactive
import matplotlib.pyplot as plt
#plt.style.use('seaborn-whitegrid')
# from pathlib import Path

##SPECIFIC
from mincalclib import mineral_constants, dataset, conversion # ,plotdata
##from mincalclib.conversion import exportAX
# from mincalclib.mainMenu import Example

#import treeviewMinLabels

'''
https://pypi.org/project/PySimpleGUI/
https://pysimplegui.readthedocs.io/tutorial/
https://opensource.com/article/18/8/pysimplegui
'''

print(sys.version)
print(sys.version_info)


class pyMin:
    version = '3.2'
    def __init__(self):
        self.version


def main():

### START using GUI or batch
## python pyMin_old.py (GUI mode)
## python pyMin_old.py inuptfile.xlsx (batch mode)

## df>> data in formato PANDAS
## usati per plot ma forse utili per rielaborazioni
## ad esempio fare export AX o EXCEL


    if len(sys.argv) >= 2:
        inputfile = sys.argv[1]
        #print('file exists')
        print()
        batch_calc_from_file(inputfile)
        exit(0)  # Successful exit
    #    quit()
    else:
        print('start GUI')


    # print("pyMin_old.py path using --> print(os.path.dirname(os.path.realpath(__file__))):\n")
    # print(os.path.dirname(os.path.realpath(__file__)))
    print("\n")
    #print("pyMin_old.py path using --> print(os.path.dirname(os.path.realpath(sys.argv[0])))\n")
    print(os.path.dirname(os.path.realpath(sys.argv[0])))
    filewd = os.path.dirname(os.path.realpath(sys.argv[0]))

    # =======GUI========================================================
    #
    root = Tk()
    root.geometry("370x520+400+30");
    #root.place(x=0, y=0, relwidth=1, relheight=1)
    # root.title("pyMin v.3 - Michele Zucali 2019")
    # label_root=Label(root, text="Mineral Formula Calculation \n\n michele.zucali@unimi.it\n\n\n "
    #                             "The Houston Rockets version - Nov-Dec 2019 \n\n\n\n\n\n\n")


    # root.title("pyMin v.3 - Michele Zucali 2021")
    # label_root = Label(root, text="Mineral Formula Calculation \n\n michele.zucali@unimi.it\n\n\n "
    #                           "The Vaccine Version - Jan 2021 \n\n\n\n\n\n\n")


    # root.title("pyMin v.3.1 - Michele Zucali 2021")
    # label_root = Label(root, text="Mineral Formula Calculation \n\n michele.zucali@unimi.it\n\n\n "
    #                          "The Lyon Version - Nov 2021 \n\n\n\n\n\n\n")

    # root.title("pyMin v.3.2 - Michele Zucali 2022")
    label_root = Label(root, text="Mineral Formula Calculation \n\n michele.zucali@unimi.it\n\n\n "
                              "The SVertical Version - Dec 2022 \n\n\n\n\n\n\n")
    label_root.pack()
##### MENU
    # app = Example()
####
    label_window = Toplevel(root)
    label_window.geometry('370x600+30+30');
    display_mineral = Label(label_window, text="Mineral Labels and Oxigens")
    display_mineral.pack()

    Lb_mineral = Listbox(label_window, height=33)
    for k in sorted(mineral_constants.mineral_oxigens.keys()):
        Lb_mineral.insert(END, (k, '->',mineral_constants.mineral_oxigens[k]))
    Lb_mineral.pack()

    output_display = Toplevel(root)
    output_display.geometry('600x200+800+30')
    label = Label(output_display, text="Output Terminal")
    label.pack()
    #text = Text(output_display)
    text = scrolledtext.ScrolledText(output_display)
    text.pack()
    text.insert(END, "Ready to RUN\n\n")
    #text.configure(state='disabled')

    root.filename = ''
    #fileOUT = ''
    #df = pd.DataFrame
    #data999={};

    def UploadAction(event=None):
        root.filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select mineral oxides input file",
                                               filetypes=[("TXT files", "*.txt"),("TAB files", "*.tab"),
                                                          ("CSV files", "*.csv"),("excel files", "*.xlsx"),
                                                          ("excel files", "*.xls")])
        print('Selected root.filename:', root.filename)
        text.insert(END, ('Selected root.filename:', root.filename))
        #return root.filename
        saved = sys.stdout
        fout = open("pyMinCalc.log", 'w')
        sys.stdout = writer(sys.stdout, fout)
        print("main in pyMin development")
        pyPTinst = pyMin()
        print("pyMin version....", format(pyPTinst.version))
        print("LOGFILE for pyMinCalc")
        print("pyMinCalc.log")
        global ds
        ds=dataset.dataset(root.filename)

        print("pandas!")
        data_pd = pd.read_excel(root.filename, engine='openpyxl')

        global fileOUT
        fileOUT=ds.return_fileout()
        #print("fileout fileout")
        #print("FILEOUT: dataset: ", type(data999))
        #global df
        #df = pd.read_excel(fileOUT, sheet_name='APPEND')

        processedFile = open('pyMinCalc.log', 'r')
        # ttk.Label(Tab4, text=[processedFile.read()]).place(x=0, y=27)
        text1 = processedFile.read()

        #text.insert(END, text1)
        text.insert(END, "\nfinishing\n")
        text.insert(END, "\ntask completed\n")
        text.insert(END, "\nsaved to file _OUT and logfile\n")
        print("finishing")
        print("executed")
        sys.stdout = saved
        fout.close()
        print("\n")
        #return data999

    def plotData(event=None):
        print("plot")
        #print(matplotlib.get_backend()) 2 novembre 21
        #pandas_data = pd.read_excel(fileOUT, sheet_name='APPEND', index_col=0)
        #print(pandas_data)
        global df
        df = pd.read_excel(fileOUT, sheet_name='APPEND', engine='openpyxl')
        print("to list")
        print(df.columns.tolist())
        df1 = df[['SiO2', 'Al2O3', 'FeO', 'MgO', 'mineral']]
        df2 = df[['SiO2', 'Na2O', 'CaO', 'K2O', 'mineral']]


        #fig, axes = plt.subplots(1,2,sharex=True,figsize=(10,5))
        #fig.suptitle('DATA PLOTS')
        #fig1 = sns.figure()

        #plt.figure(1)
        ##sns.pairplot(df1, hue="mineral",height=1.5)


        #plt.figure(2)
        ##sns.pairplot(df2, hue="mineral",height=1.4, aspect=1/1)
        #print(plt.get_current_fig_manager())

        #mngr.window.setGeometry(50,100,640,545)
        #mngr.window.setGeometry()
        #fig.canvas.manager.window.Move(100, 400)
        #plt.figure(1).show() 2 novembre 21
        #plt.figure(2).show() 2 novembre 21



    def plotDataPandasGUI(): #https://github.com/adamerose/pandasgui
        root.filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select mineral oxides input file",
                                                   filetypes=[("TXT files", "*.txt"), ("TAB files", "*.tab"),
                                                              ("CSV files", "*.csv"), ("excel files", "*.xlsx"),
                                                              ("excel files", "*.xls")])
        print("plot")
        df1 = pd.read_excel(root.filename, sheet_name='APPEND', engine='openpyxl')
        show(df1)


    def plotDataScatter():# (event=None):
        root.filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select mineral oxides input file",
                                                   filetypes=[("TXT files", "*.txt"), ("TAB files", "*.tab"),
                                                              ("CSV files", "*.csv"), ("excel files", "*.xlsx"),
                                                              ("excel files", "*.xls")])
        print("plot")
        df = pd.read_excel(root.filename, sheet_name='APPEND', engine='openpyxl')
        print("to list")
        print(df.columns.tolist())
        # df1 = df[['SiO2', 'Al2O3', 'FeO', 'MgO', 'mineral']]
        # df2 = df[['SiO2', 'Na2O', 'CaO', 'K2O', 'mineral']]

        #fig, axes = plt.subplots(figsize=(10,5),ncols=1,nrows=1)
        #fig.suptitle('DATA PLOTS')
        # fig1 = sns.figure()

        #plt.figure(1)
        #sns.pairplot(df1, hue="mineral",height=1.5)

        plt.figure(2)
        sns.scatterplot(data=df, x='SiO2', y='OxSum', hue="mineral")
        #print(plt.get_current_fig_manager())

        # mngr.window.setGeometry(50,100,640,545)
        # mngr.window.setGeometry()
        # fig.canvas.manager.window.Move(100, 400)
        #print("plot figure 1 pairplot")
        #plt.figure(1).show()# 2 novembre 21
        #plt.figure(2).show()# 2 novembre 21
        #print("plot figure 2 FacetGrid")
        #sea = sns.FacetGrid(df, col="mineral", hue="mineral")
        #sea.map(sns.scatterplot, "OxSum", "SiO2", alpha=.8)
        #sea.add_legend()
        plt.figure(2).show()


    def plotDataMatrix():# (event=None):
        root.filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select mineral oxides input file",
                                                   filetypes=[("TXT files", "*.txt"), ("TAB files", "*.tab"),
                                                              ("CSV files", "*.csv"), ("excel files", "*.xlsx"),
                                                              ("excel files", "*.xls")])
        print("plot")
        df = pd.read_excel(root.filename, sheet_name='APPEND', engine='openpyxl')
        print("to list")
        print(df.columns.tolist())
        df1 = df[['SiO2', 'Al2O3', 'FeO', 'MgO', 'mineral']]
        # df2 = df[['SiO2', 'Na2O', 'CaO', 'K2O', 'mineral']]

        #fig, axes = plt.subplots(figsize=(10,5),ncols=1,nrows=1)
        #fig.suptitle('DATA PLOTS')
        # fig1 = sns.figure()

        plt.figure(1)
        sns.pairplot(df1, hue="mineral",height=1.5)

        #plt.figure(2)
        #sns.scatterplot(df, x="SiO2", y="OxSum", hue="mineral",height=1.4, aspect=1/1)
        #print(plt.get_current_fig_manager())

        # mngr.window.setGeometry(50,100,640,545)
        # mngr.window.setGeometry()
        # fig.canvas.manager.window.Move(100, 400)
        #print("plot figure 1 pairplot")
        #plt.figure(1).show()# 2 novembre 21
        #plt.figure(2).show()# 2 novembre 21
        #print("plot figure 2 FacetGrid")
        #sea = sns.FacetGrid(df, col="mineral", hue="mineral")
        #sea.map(sns.scatterplot, "OxSum", "SiO2", alpha=.8)
        #sea.add_legend()
        plt.figure(2).show()

       ####### FINE PLOT DATA

    def convertAX(event=None):
        print("convert to AX fomatted file")
         #print(fileOUT)
        #global ds
        #global fileOUT
        fileOUT = ds.return_fileout()
        df = pd.read_excel(fileOUT, sheet_name='APPEND', engine='openpyxl')
        convert = conversion.conversioni
        convert.exportAX(df)
        #convert.exportAX(df)
        #exportAX(df)


    ###############BUTTONS
    Button(root, text="pyMin select File & Run", command=UploadAction).pack()
    Button(root).pack()
    Button(root).pack()
    Button(root, text="Plot Data PandasGUI", fg='blue', activebackground='#00ff00',
                          command=plotDataPandasGUI).pack()
    Button(root).pack()
    Button(root, text="Plot Data Matrix", fg='blue', activebackground='#00ff00',
               command=plotDataMatrix).pack()
    Button(root, text="Convert AX", fg='blue', activebackground='#00ff00', command=convertAX).pack()
    Button(root).pack()
    Button(root, text="Exit", command=root.destroy).pack()
    root.mainloop()
    # start_button = Button(root, text="Select the File & Run", command=UploadAction).pack()
    # close_button = Button(root, text="Exit", command=root.destroy).pack()
    # plot_button = Button(root, text="Plot Data Results", fg='blue',activebackground='#00ff00', command=plotData).pack()
    # ax_button = Button(root, text="convert AX", fg='blue',activebackground='#00ff00',command=convertAX).pack()
    # root.mainloop()

##### BACK TO TERMINAL-CONSOLE
    print("LIST OF LABELS and associated OXYDES\n")
    for k, v in mineral_constants.mineral_oxigens.items():
        print(k, v)

    ######## SALVA SU XLSX elenco labels per minerali


    workbookNew = openpyxl.Workbook()
    #workbookNew = xlsxwriter.Workbook(root.filename + '_labelsMIN.xlsx')
    worksheetNEW = workbookNew.worksheets[0]
    worksheetNEW.title="label"
   # worksheetNEW.name["label"]

    row = 0
    col = 1

    for key, value in mineral_constants.mineral_labels.items():
        row += 1
#        print("key, value in mineral_constants.mineral_labels.items()")
#        print(key, value)
        worksheetNEW.cell(row, col).value = key
        #       for item in mineral_constants.mineral_labels[key]:
        worksheetNEW.cell(row, col + 1).value =value
    #       row += 1
    workbookNew.save(root.filename + '_labelsMIN.xlsx')
    workbookNew.close()
    ##############


    print("system closing")
    sys.exit(0)




    '''
    TO BE implemented:
    1) cation partitioning
    2) AX export and reformat for EXCEL spreadsheets
    3) √ save each mineral
    4) SD
    5) thermobarometry?
    6) √ plot √
    7) √ remove spaces headers if any==> inoutfile.py==> line 90 in def readFILE_E_ESTRAI_DATI_MA_CONTROLLA_MINLABEL_SET_OX(file_Input_XLSX): headers.append(str(s.cell(riga,colo).value).strip())
    
    '''


class writer:
    def __init__(self, *writers):
        self.writers = writers

    def write(self, text):
        for w in self.writers:
            w.write(text)


def windowsss():
    # import PySimpleGUI as sg      

    layout = [[sg.Text('Persistent window')],
              [sg.Input(do_not_clear=True)],
              [sg.Button('Read'), sg.Exit()]]

    window = sg.Window('Window that stays open', layout)

    while True:
        event, values = window.Read()
        if event is None or event == 'Exit':
            break
        print(event, values)

    window.Close()


def batch_calc_from_file(infile):
    saved = sys.stdout
    fout = open("pyMinCalc.log", 'w')
    sys.stdout = writer(sys.stdout, fout)

    print("main in pyMin development")

    pyPTinst = pyMin()
    print("pyMin version....", format(pyPTinst.version))

    print("LOGFILE for pyMinCalc")
    print("pyMinCalc.log")

    print("RUNNING batch scripting from batch_calc_from_file()")

    dataset.dataset(infile)
    processedFile = open('pyMinCalc.log', 'r')
        # ttk.Label(Tab4, text=[processedFile.read()]).place(x=0, y=27)
    text1 = processedFile.read()

    workbookNew = openpyxl.Workbook(infile + '_labelsMIN.xlsx')
    #workbookNew = xlsxwriter.Workbook(infile + '_labelsMIN.xlsx')
    worksheetNEW = workbookNew.add_worksheet()

    row = 0
    col = 0

    for key, value in mineral_constants.mineral_labels.items():
        row += 1
        worksheetNEW.write(row, col, key)
        #       for item in mineral_constants.mineral_labels[key]:
        worksheetNEW.write(row, col + 1, value)
    #       row += 1

    workbookNew.close()

    print("finishing")
    print("executed")
    sys.stdout = saved
    fout.close()
    print("\n")


if __name__ == '__main__':
    main()
