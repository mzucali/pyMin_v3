'''
Created on Mar 28, 2019

@author: miki
'''
# import START

# from openpyxl.writer.excel import ExcelWriter


import os
import os.path
from collections import OrderedDict

import openpyxl

from mincalclib import mineral_constants
# from openpyxl import worksheet, Workbook
import pandas as pd
# pip install pandas-path
from pandas_path import path
# Use the .path accessor to get just the filename rather than the full path
# import xlwt
# import xlsxwriter
from openpyxl import load_workbook, Workbook

import xlrd
# from xlrd import open_workbook
# xlrd.xlsx.ensure_elementtree_imported(False, None)
# xlrd.xlsx.Element_has_iter = True

row_oxides = 0
col_oxides = 0


# workbook = xlsxwriter.Workbook()

def check_filein_if_xlsx(filein):
    #test1 = filein.path
    #print("str(filein) :",test1)
    print("filelin: ", filein)
    fin = filein
    filename1, file_extension1 = os.path.splitext(fin)
    print("ininoutfile fin: ", fin)
    print("ininoutfile filename1: ", filename1)
    print("ininoutfile extension: ", file_extension1)
    # extension = os.path.splitext(fin)[1]

    ## CHECK XLSX, XLS, CSV, TXT
    ext_tab = (".txt", ".tab")
    ext_csv = (".csv")
    # ext_xls = (".xls")
    # ext_xlsx = (".xlsx")

    # if fin.endswith(tuple(".xlsx")):  #it should be a folder
    if file_extension1 == ".xlsx":
        print("you're ok!")

    # elif fin.endswith(tuple(".xls")):
    elif file_extension1 == ".xls":
        print("this is an old excel file, let's convert it!")
        df = pd.read_excel(fin, header=None)
        print("df: ", df)
        df.to_excel(filename1+"_new.xlsx", index=False, header=False)
        fin = filename1+"_new.xlsx"
        print("this must be the new file: ", fin)

    elif file_extension1 == ".txt":
        # how to use the list ?
        print("sounds wierd...", fin)
        read_file = pd.read_csv(fin, sep='\t')
        # print("read file type: ",type(read_file))
        print("new file name: ", filename1 + '.xlsx')
        #  fin = filename1 + '_new.xlsx'  ##<< QUESTO PRODUCE UN OGGETTO vuoto se pero' filename1 è gia' file,
        # allora si porta dietro i dati
        print("tell me 1: ", fin, type(fin))
        read_file.to_excel(filename1 + '_new.xlsx', index=None, header=True)
        fin = filename1 + "_new.xlsx"
        print("tell me 2aa: ", fin, type(fin))
        print("I made the conversion for you.....you're welcome!")
        print("the new input file, xlsx format, is: ", fin)

    elif file_extension1 == ".tab":
        # how to use the list ?
        print("sounds wierd...", fin)
        read_file = pd.read_csv(fin, sep='\t')
        # print("read file type: ",type(read_file))
        print("new file name: ", filename1 + '.xlsx')
        #  fin = filename1 + '_new.xlsx'  ##<< QUESTO PRODUCE UN OGGETTO vuoto se pero' filename1 è gia' file,
        # allora si porta dietro i dati
        print("tell me 1: ", fin, type(fin))
        read_file.to_excel(filename1 + '_new.xlsx', index=None, header=True)
        fin = filename1 + "_new.xlsx"
        print("tell me 2aa: ", fin, type(fin))
        print("I made the conversion for you.....you're welcome!")
        print("the new input file, xlsx format, is: ", fin)

    elif file_extension1 == ".csv":
        # how to use the list ?
        print("sounds wierd...", fin)
        read_file = pd.read_csv(fin)
        # print("read file type: ",type(read_file))
        print("new file name: ", filename1 + '.xlsx')
        #  fin = filename1 + '_new.xlsx'  ##<< QUESTO PRODUCE UN OGGETTO vuoto se pero' filename1 è gia' file,
        # allora si porta dietro i dati
        print("tell me 1: ", fin, type(fin))
        read_file.to_excel(filename1 + '_new.xlsx', index=None, header=True)
        fin = filename1 + "_new.xlsx"
        print("tell me 2aa: ", fin, type(fin))
        print("I made the conversion for you.....you're welcome!")
        print("the new input file, xlsx format, is: ", fin)

    else:  # not a video, not a folder
        exit()

    print("tell me location and type of fin: ", fin, type(fin))
    return fin


def readFILE_E_ESTRAI_DATI_MA_CONTROLLA_MINLABEL_SET_OX(file_Input_XLSX):
    data_input_Ox_dict_list = []
    fin = file_Input_XLSX

    fin = check_filein_if_xlsx(fin)  # CHECK CSV, TXT, TAB, XLS (false)
    # wb = open_workbook(fin)
    # wb = pd.read_excel(fin, engine='openpyxl') ## 22 dec 22 - prova NON usare xlrd (open_workbook) ma read_excel di openpyxl
    print("inoutfile line 100")
    print(fin)
    wb = openpyxl.load_workbook(fin)
    print(wb)
    s = wb.active
    #  for s in wb.worksheets:  ##worksheets() iterate
    print('\n\nSheet: {0}, columns = {1}, rows = {2} '.format(s.title, s.max_column, s.max_row))
    for riga in range(1):

        headers = []  # global headers
        for colo in range(s.max_column):
            print("line 99: ", riga + 1, colo + 1)
            headers.append(str(s.cell(riga + 1, colo + 1).value).strip())
        # print ('headers read in Sheet: ', s.name)
        # print("SAMPLE: ",headers[0])
        # headers[0] = 'sample'
        headers[1] = 'mineral'
        # print("SAMPLE: ",headers[0])
        # print("MINERAL ", headers[1])
        print('\t'.join(i for i in headers))

    print()

    print("LIST ALL ANALYSES and STORE in dict_list")
    #  for s in wb.worksheets:
    print('\nSheet:', s.title)
    print()
    for rows in range(1, s.max_row):
        values = []
        for cols in range(s.max_column):
            values.append(s.cell(rows + 1, cols + 1).value)
        print("Analysis = %s" % rows)
        print("Headers:\t" + '\t'.join([str(i) for i in headers]))
        print("Values:\t" + '\t'.join([str(i) for i in values]))
        # print ("\n")

        data_input_Ox_dict = OrderedDict(zip(headers, values))
        print("INOUTFILE data_input_Ox_dict = ", data_input_Ox_dict)
        #             for k,v in data_input_Ox_dict.items():
        #                 print("k e v: ", k, v)
        data_input_Ox_dict2 = OrderedDict(zip(headers, values))
        data_input_Ox_dict2 = changeMineralLabels(data_input_Ox_dict)
        print("INOUTFILE data_input_Ox_dict CHANGED MIN LABEL= ", data_input_Ox_dict2)

        # data_input_Ox_dict_list = []
        # print("from READEXCEL2 added analysis = ",data_input_Ox_dict)
        # print()
        data_input_Ox_dict_list.append(data_input_Ox_dict2)

    print()

    # wb.release_resources() ## xlrd

    ## SET mineral to 'mineral' and sample to 'Sample'
    ## CHECK MINERAL LABEL and change to unique LABEL=>>DONE
    ## 

    print("INOUTdata_input_Ox_dict_list..", data_input_Ox_dict_list)

    return data_input_Ox_dict_list


def changeMineralLabels(cats_dict):
    if cats_dict['mineral'] in mineral_constants.mineral_labels.keys():
        v = cats_dict['mineral']
        print("changeALLKeys found: ", v)

        print(mineral_constants.mineral_labels[v])
        new_value = mineral_constants.mineral_labels[v]
        cats_dict['mineral'] = new_value

    return cats_dict


def removeEXT(file_name_path):
    filename = os.path.splitext(file_name_path)[0]
    return filename


def write_out_base_data(recalc_data_oxides_cats_OX__list, inputfile_path):
    global wbook
    global wsheet

    print("FILE INPUT ..." + inputfile_path)
    file_output = removeEXT(inputfile_path) + '_OUT.xlsx'
    print("FILE OUTPUT ..." + file_output)
    print("WRITING HEADERS to FileOUT " + str(file_output))

    # wbook = openpyxl.Workbook(file_output)
    wbook = openpyxl.Workbook()
    wsheet = wbook.create_sheet("data")

    row = 1
    col = 1
    print("DATASET write headers from oxides_dict_list...", recalc_data_oxides_cats_OX__list)
    print("DATASET write headers from self.data_input_OX_list", recalc_data_oxides_cats_OX__list)
    for k in recalc_data_oxides_cats_OX__list[0]:
        print("k=> ", str(k))

        #       wsheet.write(row, col, k)
        wsheet.cell(row, col).value = k
        row += 1
    row = 1
    col += 1

    print("column = " + str(col))
    print("row = " + str(row))

    # wbook.save(file_output)

    #     global headers_written
    #     headers_written = True

    global row_oxides
    row_oxides = row
    global col_oxides
    col_oxides = col

    for mineral in recalc_data_oxides_cats_OX__list:
        for k, v in mineral.items():
            wsheet.cell(row, col).value = v
            row += 1
        col += 1
        row = 1

    print("GLOBAL WNOOK: ", wbook)

    wbook.save(file_output)
    transpose_excel(file_output)

    wbook.close()
    return file_output


def transpose_excel(fileOUT):
    df = pd.read_excel(fileOUT, sheet_name="data", engine='openpyxl')
    # df.as_matrix()
    df.values
    df.transpose()
    print("table")
    print(str(df))
    print("transpose")
    print(str(df.transpose()))
    writer1 = pd.ExcelWriter(fileOUT)
    df.to_excel(writer1, sheet_name='TAB', index=False)
    df.transpose().to_excel(writer1, sheet_name='APPEND', header=False)
    writer1.save()

    return


def write_out_data_by_mineral_with_specific_sites2(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    global wb2
    wb2 = load_workbook(fileOUT)
    # print("WB2 = ", wb2)

    for k, v in recalc_data_oxides_cats_OX_by_mineral_list.items():
        # print("k and v: ", k, v)
        # print("WB2 = ", wb2)

        global sheet
        sheet = wb2.create_sheet(k)

        global row
        global col
        row = 1
        col = 1

        for l in v[0]:
            sheet.cell(row=row, column=col).value = l
            row += 1

        row = 1
        col += 1

        for mineral in v:
            for k, v in mineral.items():
                sheet.cell(row=row, column=col).value = v
                row += 1
            col += 1
            row = 1

    wb2.save(fileOUT)


def write_out_data_by_mineral_with_specific_sites(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    print("INOUTFILE=>write_out_data_by_mineral_with_specific_sites(recalc_data_oxides_cats_OX_by_mineral_list)")

    for k, v in recalc_data_oxides_cats_OX_by_mineral_list.items():
        print("k and v: ", k, v)

        print("FILE INPUT ..." + fileOUT)
        file_output = removeEXT(fileOUT) + '_' + k + '.xlsx'
        print("FILE OUTPUT ..." + file_output)

        wbook = openpyxl.Workbook(file_output)
        wsheet = wbook.create_sheet(k)

        global row
        global col
        row = 0
        col = 0

        for l in v[0]:
            wsheet.write(row, col, l)
            row += 1
        row = 0
        col += 1

        for mineral in v:
            for k, v in mineral.items():
                wsheet.write(row, col, v)
                row += 1
            col += 1
            row = 0

        wbook.close()

    return


def writeAX_formatted_input(recalc_data_oxides_cats_OX_by_mineral_list, fileOUT):
    ##to be implemented
    '''
    write headers
    read by mineral
    write by mineral
        sample mineral (check AX)
        new line
        write oxides
    
    '''
    return
